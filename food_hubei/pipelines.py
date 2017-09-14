# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: http://doc.scrapy.org/en/latest/topics/item-pipeline.html

import time
from openpyxl import Workbook, load_workbook
import os

global sheet_index

class FoodTypePipeline(object):
    def process_item(self, item, spider):
        today = time.strftime('%Y%m%d', time.localtime())
        file_name = today + 'food_type.txt'
        with open(file_name, 'a') as f:
            for key in item:
                f.write('%s: %s\n' % (key.encode('utf-8'),
                                      item[key].encode('utf-8'))
                        )
            f.write(3 * '\n')
        return item        

class FoodHubeiPipeline(object):
    def process_item(self, item, spider):
        today = time.strftime('%Y%m%d', time.localtime())
        file_name = today + 'food_hubei.txt'
        with open(file_name, 'a') as f:
            for key in item:
                f.write('%s: %s\n' % (key,
                                      unicode(item[key]).encode('utf-8'))
                        )
            f.write(3 * '\n')
        return item
    
class FoodExcelPipeline(object):
    def process_item(self, item, spider):
        if not os.path.exists(u'湖北食品检查.xlsx'):
            wb = Workbook()
            wb.save(u'湖北食品检查.xlsx')
            
        wb = load_workbook(u'湖北食品检查.xlsx')
        
        try:
            ws = wb.get_sheet_by_name(item['food_type'])
        except:
            index = len(wb.get_sheet_names())
            ws = wb.create_sheet(title=item['food_type'], index=index-1)
            i = 65
            for key in item:
                ws[chr(i)+'1'] = key
                print chr(i)+'1'
                i += 1
        
        last_row = ws.max_row+1
        for i in range(1, 14):
            ws.cell(row=last_row, column=i).value = item[ws.cell(row=1, column=i).value]
            
        wb.save(u'湖北食品检查.xlsx')
        return item
